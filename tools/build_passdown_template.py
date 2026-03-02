"""Build a standardized passdown Excel template that preserves the familiar block layout.

Structure matches the existing passdown supervisors already use:
  - Each block = one shift × line × product run
  - Header rows: Product, Order#, Cases, OEE (same positions as current)
  - Downtime rows: sequence #, Area (dropdown), Failure Mode (dropdown), details, etc.
  - Dropdowns enforce consistency without changing the feel

Produces:
  - "Passdown" sheet with pre-built blocks for 5 lines × 1 shift
  - "Reference" sheet (hidden) powering dropdowns
  - "Rates" sheet with targets
  - "Example" sheet showing a filled-in block
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SCHEMA_PATH = Path(__file__).resolve().parent.parent / "memory" / "passdown_schema.json"
KB_PATH = Path(__file__).resolve().parent.parent / "memory" / "equipment_kb.json"

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Calibri", size=10, bold=True)
BLOCK_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DT_HEADER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium"))


def _build_reference_sheet(wb, schema):
    """Build hidden Reference sheet with dropdown values."""
    ref = wb.create_sheet("Reference")
    ref.sheet_state = "hidden"
    ref_map = {}
    col = 1

    # Equipment areas
    areas = schema["equipment_areas"]["values"]
    ref.cell(row=1, column=col, value="Equipment").font = Font(bold=True)
    for i, v in enumerate(areas, 2):
        ref.cell(row=i, column=col, value=v)
    ref_map["Equipment"] = f"Reference!${get_column_letter(col)}$2:${get_column_letter(col)}${len(areas)+1}"
    col += 1

    # Non-equipment
    non_equip = schema["non_equipment_categories"]["values"]
    ref.cell(row=1, column=col, value="Non-Equipment").font = Font(bold=True)
    for i, v in enumerate(non_equip, 2):
        ref.cell(row=i, column=col, value=v)
    col += 1

    # Combined areas (equipment + non-equipment)
    all_areas = areas + non_equip
    ref.cell(row=1, column=col, value="All Areas").font = Font(bold=True)
    for i, v in enumerate(all_areas, 2):
        ref.cell(row=i, column=col, value=v)
    ref_map["All Areas"] = f"Reference!${get_column_letter(col)}$2:${get_column_letter(col)}${len(all_areas)+1}"
    col += 1

    # Failure modes
    modes = schema["failure_modes"]["values"]
    ref.cell(row=1, column=col, value="Failure Mode").font = Font(bold=True)
    for i, v in enumerate(modes, 2):
        ref.cell(row=i, column=col, value=v)
    ref_map["Failure Mode"] = f"Reference!${get_column_letter(col)}$2:${get_column_letter(col)}${len(modes)+1}"
    col += 1

    # Shifts
    ref.cell(row=1, column=col, value="Shifts").font = Font(bold=True)
    for i, v in enumerate(schema["shifts"], 2):
        ref.cell(row=i, column=col, value=v)
    ref_map["Shifts"] = f"Reference!${get_column_letter(col)}$2:${get_column_letter(col)}${len(schema['shifts'])+1}"
    col += 1

    # Lines
    ref.cell(row=1, column=col, value="Lines").font = Font(bold=True)
    for i, v in enumerate(schema["lines"], 2):
        ref.cell(row=i, column=col, value=v)
    ref_map["Lines"] = f"Reference!${get_column_letter(col)}$2:${get_column_letter(col)}${len(schema['lines'])+1}"
    col += 1

    # Resolved
    ref.cell(row=1, column=col, value="Resolved").font = Font(bold=True)
    ref.cell(row=2, column=col, value="Yes")
    ref.cell(row=3, column=col, value="No")
    ref_map["Resolved"] = f"Reference!${get_column_letter(col)}$2:${get_column_letter(col)}$3"

    return ref_map


def _write_block(ws, row, ref_map, dv_cache, line_num=None, example=None):
    """Write one production block (header + downtime rows). Returns next available row.

    Block layout (mirrors existing passdown):
      Row 0: [blank] | "Product:" | [free text] | Shift ▼ | Line ▼ | Date
      Row 1: [blank] | "Order #:" | [free text] | (shift/line/date carry down)
      Row 2: [blank] | "Cases:"   | [number]    |
      Row 3: [blank] | "OEE:"     | [pct]       |
      Row 4: [blank] | [blank]    | [blank]     | [blank] | [blank] | [blank] | # | Area ▼ | Issue/Details | Failure Mode ▼ | Action | Result/Resolved ▼ | Duration(min) | Notes
      Rows 5-12: 8 downtime entry rows

    Columns:
      A: (spacer)
      B: Labels / sequence #
      C: Values (product, order, cases, OEE) / issue details  
      D: Shift
      E: Line
      F: Date
      G: Downtime # (sequence)
      H: Area (dropdown)
      I: Issue / Details (free text)
      J: Failure Mode (dropdown)  
      K: Action Taken
      L: Resolved? (dropdown)
      M: Duration (min)
      N: Notes
    """
    r = row

    # ── Production header rows ──
    for col_idx in range(1, 15):
        ws.cell(row=r, column=col_idx).fill = BLOCK_HEADER_FILL
        ws.cell(row=r, column=col_idx).border = THIN_BORDER

    ws.cell(row=r, column=2, value="Product:").font = LABEL_FONT
    ws.cell(row=r, column=2).fill = BLOCK_HEADER_FILL
    ws.cell(row=r, column=3).fill = BLOCK_HEADER_FILL  # product name entry
    # Shift dropdown
    _add_validation(ws, ws.cell(row=r, column=4), ref_map["Shifts"], dv_cache)
    ws.cell(row=r, column=4).fill = BLOCK_HEADER_FILL
    # Line dropdown
    _add_validation(ws, ws.cell(row=r, column=5), ref_map["Lines"], dv_cache)
    ws.cell(row=r, column=5).fill = BLOCK_HEADER_FILL
    # Date
    ws.cell(row=r, column=6).number_format = "YYYY-MM-DD"
    ws.cell(row=r, column=6).fill = BLOCK_HEADER_FILL

    if example:
        ws.cell(row=r, column=3, value=example.get("product", ""))
        ws.cell(row=r, column=4, value=example.get("shift", ""))
        ws.cell(row=r, column=5, value=example.get("line", ""))
        ws.cell(row=r, column=6, value=example.get("date", ""))

    r += 1

    # Order #
    for col_idx in range(1, 15):
        ws.cell(row=r, column=col_idx).fill = BLOCK_HEADER_FILL
        ws.cell(row=r, column=col_idx).border = THIN_BORDER
    ws.cell(row=r, column=2, value="Order #:").font = LABEL_FONT
    if example:
        ws.cell(row=r, column=3, value=example.get("order", ""))
    r += 1

    # Cases
    for col_idx in range(1, 15):
        ws.cell(row=r, column=col_idx).fill = BLOCK_HEADER_FILL
        ws.cell(row=r, column=col_idx).border = THIN_BORDER
    ws.cell(row=r, column=2, value="Cases:").font = LABEL_FONT
    if example:
        ws.cell(row=r, column=3, value=example.get("cases", ""))
    r += 1

    # OEE
    for col_idx in range(1, 15):
        ws.cell(row=r, column=col_idx).fill = BLOCK_HEADER_FILL
        ws.cell(row=r, column=col_idx).border = THIN_BORDER
    ws.cell(row=r, column=2, value="OEE:").font = LABEL_FONT
    ws.cell(row=r, column=3).number_format = "0.0%"
    if example:
        ws.cell(row=r, column=3, value=example.get("oee", ""))
    r += 1

    # ── Downtime header row ──
    dt_headers = {
        7: "#",
        8: "Area ▼",
        9: "Issue / Details",
        10: "Failure Mode ▼",
        11: "Action Taken",
        12: "Resolved? ▼",
        13: "Time(min)",
        14: "Notes",
    }
    for col_idx in range(1, 15):
        ws.cell(row=r, column=col_idx).fill = DT_HEADER_FILL
        ws.cell(row=r, column=col_idx).border = THIN_BORDER
    for col_idx, label in dt_headers.items():
        cell = ws.cell(row=r, column=col_idx, value=label)
        cell.font = Font(name="Calibri", size=9, bold=True)
        cell.fill = DT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    r += 1

    # ── Downtime entry rows (8 per block) ──
    num_dt_rows = 8
    for dt_idx in range(1, num_dt_rows + 1):
        for col_idx in range(1, 15):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.cell(row=r, column=7, value=dt_idx)  # sequence number
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center")

        # Area dropdown
        _add_validation(ws, ws.cell(row=r, column=8), ref_map["All Areas"], dv_cache)
        # Failure Mode dropdown
        _add_validation(ws, ws.cell(row=r, column=10), ref_map["Failure Mode"], dv_cache)
        # Resolved dropdown
        _add_validation(ws, ws.cell(row=r, column=12), ref_map["Resolved"], dv_cache)

        # Fill example data
        if example and dt_idx <= len(example.get("events", [])):
            evt = example["events"][dt_idx - 1]
            ws.cell(row=r, column=8, value=evt.get("area", ""))
            ws.cell(row=r, column=9, value=evt.get("issue", ""))
            ws.cell(row=r, column=10, value=evt.get("failure_mode", ""))
            ws.cell(row=r, column=11, value=evt.get("action", ""))
            ws.cell(row=r, column=12, value=evt.get("resolved", ""))
            ws.cell(row=r, column=13, value=evt.get("duration", ""))
            ws.cell(row=r, column=14, value=evt.get("notes", ""))

        r += 1

    # Spacer row
    r += 1
    return r


def _add_validation(ws, cell, formula, dv_cache):
    """Add data validation, reusing DV objects (Excel limits to 255 per sheet)."""
    if formula not in dv_cache:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.showDropDown = False  # False = show the dropdown arrow (confusing API)
        ws.add_data_validation(dv)
        dv_cache[formula] = dv
    dv_cache[formula].add(cell)


def build_template(output_path: str | Path, schema_path: str | Path | None = None) -> Path:
    """Build the passdown Excel template."""
    schema = json.loads((Path(schema_path) if schema_path else SCHEMA_PATH).read_text())
    output_path = Path(output_path)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Reference sheet
    ref_map = _build_reference_sheet(wb, schema)

    # ── Main Passdown sheet ──
    ps = wb.create_sheet("Passdown", 0)

    # Column widths
    widths = {
        "A": 2, "B": 12, "C": 25, "D": 8, "E": 6, "F": 12,
        "G": 4, "H": 22, "I": 40, "J": 20, "K": 35, "L": 10,
        "M": 10, "N": 30,
    }
    for col, w in widths.items():
        ps.column_dimensions[col].width = w

    # Title
    ps.merge_cells("A1:N1")
    title = ps.cell(row=1, column=1, value="Production Passdown — Shift Handoff Log")
    title.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    title.alignment = Alignment(horizontal="center")

    ps.merge_cells("A2:N2")
    instr = ps.cell(row=2, column=1,
                    value="One block per line/product run. Fill production info (blue), "
                          "then log each downtime event below (yellow). Use dropdowns for Area, Failure Mode, and Resolved.")
    instr.font = Font(name="Calibri", size=9, italic=True, color="666666")

    # Write blocks for 5 lines
    row = 4
    dv_cache: dict[str, DataValidation] = {}
    for line in range(1, 6):
        row = _write_block(ps, row, ref_map, dv_cache, line_num=line)

    # Add 5 more blank blocks for additional runs
    for _ in range(5):
        row = _write_block(ps, row, ref_map, dv_cache)

    ps.freeze_panes = "A4"

    # ── Example sheet ──
    ex = wb.create_sheet("Example")
    for col, w in widths.items():
        ex.column_dimensions[col].width = w

    ex.merge_cells("A1:N1")
    ex.cell(row=1, column=1, value="Example — Filled Passdown Block").font = Font(
        name="Calibri", size=14, bold=True, color="1F4E79")

    ex_dv_cache: dict[str, DataValidation] = {}
    _write_block(ex, 3, ref_map, ex_dv_cache, example={
        "product": "DM WK GOLD CORN",
        "shift": "3rd",
        "line": 1,
        "date": "2026-03-01",
        "order": "1184500L1",
        "cases": 2500,
        "oee": 0.22,
        "events": [
            {
                "area": "Bear Labeler",
                "issue": "Loose labels, glue not adhering to cans consistently",
                "failure_mode": "Loose / Misapplied Labels",
                "action": "Operator cleaned curling bar and adjusted fingers, called maintenance",
                "resolved": "No",
                "duration": 45,
                "notes": "",
            },
            {
                "area": "Tray Packer / Caser",
                "issue": "Can jams at infeed, cases double-feeding",
                "failure_mode": "Jam / Blockage",
                "action": "Operator cleared jams, maintenance adjusted lane guides",
                "resolved": "Yes",
                "duration": 30,
                "notes": "",
            },
            {
                "area": "Lunch & Breaks",
                "issue": "No relief operators available",
                "failure_mode": "",
                "action": "",
                "resolved": "",
                "duration": 50,
                "notes": "Stopped for all breaks",
            },
        ],
    })

    # ── Rates sheet ──
    rates = wb.create_sheet("Rates")
    rate_headers = ["Line", "Product", "Cases/Shift", "Cases/Pallet", "Cans/Case"]
    for col_idx, h in enumerate(rate_headers, 1):
        cell = rates.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        rates.column_dimensions[get_column_letter(col_idx)].width = 14

    if KB_PATH.is_file():
        try:
            kb = json.loads(KB_PATH.read_text())
            for i, r in enumerate(kb.get("rates", []), 2):
                rates.cell(row=i, column=1, value=r.get("line", ""))
                rates.cell(row=i, column=2, value=r.get("product", ""))
                rates.cell(row=i, column=3, value=r.get("cases_per_shift"))
                rates.cell(row=i, column=4, value=r.get("cases_per_pallet"))
                rates.cell(row=i, column=5, value=r.get("cans_per_case"))
        except Exception:
            pass

    wb.save(str(output_path))
    return output_path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "Passdown_Template.xlsx"
    schema = sys.argv[2] if len(sys.argv) > 2 else None
    path = build_template(out, schema)
    print(f"Template saved: {path}")
