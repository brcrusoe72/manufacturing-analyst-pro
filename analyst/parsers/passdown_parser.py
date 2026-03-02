"""Parse supervisor passdown Excel workbook into structured JSON.

Handles two formats:
  - Early format (Dec 2025): columns = Shift, Line, Date, #, Area, Details, Time(min), Notes
  - Later format (Dec 16+ 2025): columns = Shift, Line, Date, #, Area, ISSUE, ACTION, RESULT, ROOT CAUSE/RESOLVED, Time(min), Notes

Each daily sheet has multiple "blocks" — one per line per product run.
Block header rows contain Product, Order#, Cases, OEE in col B/C.
Downtime rows have a sequence number in col F (1,2,3...), Area in col G, details in col H+.

Also extracts:
  - Reference sheet (equipment list, shift defs, lines)
  - Rates sheet (target cases/shift per line/product)
  - Staffing sheets
"""
from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(slots=True)
class PassdownEvent:
    date: str                    # YYYY-MM-DD
    shift: str                   # 1st, 2nd, 3rd
    line: int | str              # 1-5
    product: str
    order_number: str
    cases: int | None
    oee: float | None
    seq: int                     # downtime sequence within block
    area: str                    # equipment area
    issue: str                   # problem description
    action: str                  # what was done
    result: str                  # outcome
    root_cause: str              # root cause or resolved/open status
    duration_minutes: int | None # Time(min)
    notes: str


@dataclass(slots=True)
class ProductionBlock:
    date: str
    shift: str
    line: int | str
    product: str
    order_number: str
    cases: int | None
    oee: float | None
    events: list[PassdownEvent] = field(default_factory=list)


@dataclass(slots=True)
class LineRate:
    line: str
    product: str
    cases_per_shift: int | None
    cases_per_pallet: int | None
    cans_per_case: int | None


@dataclass(slots=True)
class PassdownData:
    blocks: list[ProductionBlock]
    rates: list[LineRate]
    date_range: tuple[str, str]
    total_events: int
    sheets_parsed: int


# Sheets to skip (not daily passdown data)
SKIP_SHEETS = {
    'reference', 'reference ', 'template', 'template (2)', 'new_format_template',
    'new_format_template (2)', 'damaged bar', 'broken_knife_assembly',
    'jammed_cases', 'quality', 'plexi_glass', 'line details',
    'damaged cans and label', 'jammed on conveyor', 'bad cases',
    'videojet health', 'db line 2', 'rates',
}

# Sheets that are staffing data
STAFFING_PREFIX = 'staffing'


def _parse_date_from_sheet_name(name: str) -> date | None:
    """Try to parse a date from sheet name like '12-3-25', '01-01-26', '1-04-26'."""
    name = name.strip()
    # Try MM-DD-YY patterns
    for fmt in ('%m-%d-%y', '%m-%d-%Y'):
        try:
            return datetime.strptime(name, fmt).date()
        except ValueError:
            pass
    return None


def _safe_str(val: Any) -> str:
    if val is None:
        return ''
    if isinstance(val, str):
        return val.strip()
    return str(val).strip()


def _safe_int(val: Any) -> int | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def _safe_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().rstrip('%'))
    except (ValueError, TypeError):
        return None


def _safe_line(val: Any) -> int | str:
    """Extract line number."""
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        m = re.search(r'\d+', val)
        if m:
            return int(m.group())
    return str(val) if val else ''


def _extract_date_from_row(row: tuple, fallback_date: date | None) -> str:
    """Try to find a datetime in the row's date column (index 5), fall back to sheet name date."""
    val = row[5] if len(row) > 5 else None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if fallback_date:
        return fallback_date.strftime('%Y-%m-%d')
    return ''


def _detect_format(rows: list[tuple]) -> str:
    """Detect whether this sheet uses early or late format based on header row."""
    if len(rows) < 2:
        return 'early'
    header = rows[1]  # row index 1 is the header
    for cell in header:
        if isinstance(cell, str) and 'ISSUE' in cell.upper():
            return 'late'
    return 'early'


def _parse_daily_sheet(
    ws, sheet_name: str, fallback_date: date | None
) -> list[ProductionBlock]:
    """Parse one daily sheet into production blocks."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    fmt = _detect_format(rows)
    blocks: list[ProductionBlock] = []
    current_block: ProductionBlock | None = None

    for row_idx, row in enumerate(rows):
        if len(row) < 8:
            continue

        col_b = _safe_str(row[1])  # Column B — labels like Product:, Cases:, OEE:

        # Detect block header: "Product:" in col B
        if col_b.lower().startswith('product:'):
            product_name = _safe_str(row[2])
            shift = _safe_str(row[3])
            line = _safe_line(row[4])
            dt_str = _extract_date_from_row(row, fallback_date)

            if not product_name and not shift:
                # Template row, skip
                continue

            current_block = ProductionBlock(
                date=dt_str,
                shift=shift or '3rd',
                line=line,
                product=product_name,
                order_number='',
                cases=None,
                oee=None,
            )
            blocks.append(current_block)

            # Check if this row also has a downtime event (col F has seq#, col G has Area)
            seq = _safe_int(row[6])
            area = _safe_str(row[7])
            if seq and area and current_block:
                evt = _build_event(row, fmt, current_block, seq, fallback_date)
                if evt:
                    current_block.events.append(evt)
            continue

        # Order number row
        if col_b.lower().startswith('order') or col_b.lower().startswith('finished'):
            if current_block:
                current_block.order_number = _safe_str(row[2])
            # Also check for downtime event on this row
            seq = _safe_int(row[6])
            area = _safe_str(row[7])
            if seq and area and current_block:
                evt = _build_event(row, fmt, current_block, seq, fallback_date)
                if evt:
                    current_block.events.append(evt)
            continue

        # Cases row
        if col_b.lower().startswith('cases:'):
            if current_block:
                current_block.cases = _safe_int(row[2])
            seq = _safe_int(row[6])
            area = _safe_str(row[7])
            if seq and area and current_block:
                evt = _build_event(row, fmt, current_block, seq, fallback_date)
                if evt:
                    current_block.events.append(evt)
            continue

        # OEE row
        if col_b.lower().startswith('oee:'):
            if current_block:
                oee_val = _safe_float(row[2])
                if oee_val is not None and oee_val <= 1.0:
                    current_block.oee = oee_val
                elif oee_val is not None:
                    current_block.oee = oee_val / 100.0
            seq = _safe_int(row[6])
            area = _safe_str(row[7])
            if seq and area and current_block:
                evt = _build_event(row, fmt, current_block, seq, fallback_date)
                if evt:
                    current_block.events.append(evt)
            continue

        # Misc or continuation rows — check for downtime events
        seq = _safe_int(row[6])
        area = _safe_str(row[7])
        if seq and area and current_block:
            evt = _build_event(row, fmt, current_block, seq, fallback_date)
            if evt:
                current_block.events.append(evt)

    return blocks


def _build_event(
    row: tuple, fmt: str, block: ProductionBlock, seq: int,
    fallback_date: date | None
) -> PassdownEvent | None:
    """Build a PassdownEvent from a row."""
    area = _safe_str(row[7])
    if not area:
        return None

    if fmt == 'early':
        # Early format: col H = Details, col N = Time(min), col O = Notes
        issue = _safe_str(row[8])
        action = ''
        result = ''
        root_cause = ''
        duration = _safe_int(row[14]) if len(row) > 14 else None
        notes = _safe_str(row[15]) if len(row) > 15 else ''
    else:
        # Late format: col H = ISSUE, col N = ACTION, col O = RESULT, col P = ROOT CAUSE, col Q = Time(min), col R = Notes
        issue = _safe_str(row[8])
        action = _safe_str(row[14]) if len(row) > 14 else ''
        result = _safe_str(row[15]) if len(row) > 15 else ''
        root_cause = _safe_str(row[16]) if len(row) > 16 else ''
        duration = _safe_int(row[17]) if len(row) > 17 else None
        notes = _safe_str(row[18]) if len(row) > 18 else ''

    if not issue and not action:
        return None

    return PassdownEvent(
        date=block.date,
        shift=block.shift,
        line=block.line,
        product=block.product,
        order_number=block.order_number,
        cases=block.cases,
        oee=block.oee,
        seq=seq,
        area=area,
        issue=issue,
        action=action,
        result=result,
        root_cause=root_cause,
        duration_minutes=duration,
        notes=notes,
    )


def _parse_rates(ws) -> list[LineRate]:
    """Parse the Rates sheet."""
    rates = []
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:  # skip header
        if len(row) < 4:
            continue
        line = _safe_str(row[0])
        product = _safe_str(row[1])
        if not line or not product:
            continue
        rates.append(LineRate(
            line=line,
            product=product,
            cases_per_shift=_safe_int(row[2]),
            cases_per_pallet=_safe_int(row[3]),
            cans_per_case=_safe_int(row[5]) if len(row) > 5 else None,
        ))
    return rates


def parse_passdown(file_path: str | Path) -> PassdownData:
    """Parse the full passdown workbook and return structured data."""
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

    all_blocks: list[ProductionBlock] = []
    rates: list[LineRate] = []
    sheets_parsed = 0

    for name in wb.sheetnames:
        lower = name.strip().lower()

        # Skip non-data sheets
        if lower in SKIP_SHEETS:
            if lower == 'rates':
                rates = _parse_rates(wb[name])
            continue

        # Skip staffing sheets
        if lower.startswith(STAFFING_PREFIX):
            continue

        # Try to parse as daily sheet
        fallback_date = _parse_date_from_sheet_name(name)
        if fallback_date is None:
            # Not a date-named sheet, skip
            continue

        blocks = _parse_daily_sheet(wb[name], name, fallback_date)
        if blocks:
            all_blocks.extend(blocks)
            sheets_parsed += 1

    wb.close()

    # Compute stats
    total_events = sum(len(b.events) for b in all_blocks)
    dates = [b.date for b in all_blocks if b.date]
    date_range = (min(dates), max(dates)) if dates else ('', '')

    return PassdownData(
        blocks=all_blocks,
        rates=rates,
        date_range=date_range,
        total_events=total_events,
        sheets_parsed=sheets_parsed,
    )


def passdown_to_json(data: PassdownData) -> dict:
    """Convert PassdownData to a JSON-serializable dict."""
    return {
        'meta': {
            'date_range': list(data.date_range),
            'total_events': data.total_events,
            'sheets_parsed': data.sheets_parsed,
            'total_blocks': len(data.blocks),
        },
        'rates': [asdict(r) for r in data.rates],
        'blocks': [
            {
                'date': b.date,
                'shift': b.shift,
                'line': b.line,
                'product': b.product,
                'order_number': b.order_number,
                'cases': b.cases,
                'oee': b.oee,
                'events': [asdict(e) for e in b.events],
            }
            for b in data.blocks
        ],
    }


def save_passdown_json(file_path: str | Path, output_path: str | Path | None = None) -> Path:
    """Parse passdown Excel and save as JSON. Returns output path."""
    data = parse_passdown(file_path)
    js = passdown_to_json(data)

    if output_path is None:
        output_path = Path(file_path).with_suffix('.json')
    else:
        output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(js, indent=2, ensure_ascii=False))

    return output_path


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print(f"Usage: python -m analyst.parsers.passdown_parser <passdown.xlsx> [output.json]")
        sys.exit(1)

    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else None
    out = save_passdown_json(src, dst)

    # Print summary
    data = parse_passdown(src)
    print(f"Parsed {data.sheets_parsed} daily sheets")
    print(f"Date range: {data.date_range[0]} to {data.date_range[1]}")
    print(f"Production blocks: {len(data.blocks)}")
    print(f"Downtime events: {data.total_events}")
    print(f"Rates: {len(data.rates)} entries")
    print(f"Saved to: {out}")
