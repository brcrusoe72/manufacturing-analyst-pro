"""Downtime event Excel export parser."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from ._compat import ParseError, get_logger, safe_cell_value
from ._parser_utils import _cell, _infer_line_id_from_filename, _to_datetime, _to_float, _to_int, _to_text
from .oee_parser import normalize_line_id
from .utils import get_cache_dir, read_cache, write_cache

CACHE_VERSION = 7
_log = get_logger("connectors.parsers.event")

EQUIPMENT_NORMALIZATION: dict[str, str] = {
    "Palletizer - Alvey": "palletizer",
    "Labeler - Bear": "labeler",
    "Depal - Wallon": "depalletizer",
    "Tray Packer - Kayak": "tray_packer",
    "Caser Disch Rail Jam": "caser",
    "Caser Low Air Press": "caser",
    "Caser No Prod at Loader": "caser",
    "Caser Safety Interlock": "caser",
    "Caser Tipped Product": "caser",
    "Wrapper Exit Conv Busy": "wrapper",
    "Wrapper Film Out Fault": "wrapper",
    "Wrapper General Fault": "wrapper",
    "Wrapper Hot Wire Fault": "wrapper",
    "Pallet Wrapper - Highlight": "pallet_wrapper",
    "Shrink Tunnel - Kayat": "shrink_tunnel",
    "Spiral - Ryson": "spiral_conveyor",
    "Can Printer - Linx": "can_printer",
    "Tag Printer - Diagraph": "tag_printer",
    "Tray Printer - VIAcode": "tray_printer",
    "Print and Apply - Barcode Printer": "barcode_printer",
    "Vision System - TensorID": "vision_system",
    "X-Ray - Inspec": "xray",
    "Label Reader": "label_reader",
    "Can Conveyor": "can_conveyor",
    "Case Conveyor": "case_conveyor",
    "Depal Conveyor": "depal_conveyor",
}

OPERATIONAL_CATEGORIES: set[str] = {
    "Change Over",
    "Short Stop",
    "Break-Lunch",
    "Breaks/Lunch/Meals",
    "Breaks, Lunch, Meals",
    "Break Relief Other Line",
    "No Stock",
    "Bad Stock",
    "Dented Cans",
    "Greasy Cans",
    "Rusty Cans",
    "Leakers",
    "Shiners",
    "Quality",
    "Drive Off",
    "Can Codes",
    "Date Code Change",
    "Power Outage",
    "Holiday",
    "Not Scheduled",
    "Training - Meeting",
    "Meetings",
    "Other",
    "Unassigned",
    "Conveyor Problem",
    "Broken Curling Bar",
}

_EQUIPMENT_NORMALIZATION_OVERRIDE: dict[str, str] | None = None
_OPERATIONAL_CATEGORIES_OVERRIDE: set[str] | None = None


def configure_event_parser(
    *,
    equipment_normalization: dict[str, str] | None = None,
    operational_categories: set[str] | None = None,
) -> None:
    """Optionally override plant-specific normalization dictionaries."""
    global _EQUIPMENT_NORMALIZATION_OVERRIDE, _OPERATIONAL_CATEGORIES_OVERRIDE
    _EQUIPMENT_NORMALIZATION_OVERRIDE = dict(equipment_normalization) if equipment_normalization else None
    _OPERATIONAL_CATEGORIES_OVERRIDE = set(operational_categories) if operational_categories else None


@dataclass(slots=True)
class DowntimeEvent:
    """Typed downtime/performance event extracted from MES event export."""

    event_id: int
    start_time: datetime
    end_time: datetime
    duration_seconds: float
    line_id: str
    line_raw_name: str
    equipment_id: str | None
    equipment_raw_name: str
    event_type: str
    loss_type: str
    is_equipment_fault: bool
    notes: str | None


def parse_event_file(path: str | Path) -> list[DowntimeEvent]:
    """Parse Event Overview export with robust handling and cache support."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for MES Excel parsing. Install with: pip install openpyxl"
        ) from exc

    source = Path(path)
    fallback_line_id = _infer_line_id_from_filename(source.name)
    cache = get_cache_dir() / f"{source.stem.replace(' ', '_').lower()}_events.json"
    source_mtime = source.stat().st_mtime
    cached = _load_cache(cache, source_mtime)
    if cached is not None:
        _log.debug("cache hit path=%s records=%d", source.name, len(cached))
        return cached
    _log.debug("cache miss path=%s", source.name)

    wb = load_workbook(filename=source, read_only=True, data_only=True)
    try:
        for warning in validate_event_workbook(wb, str(source)):
            _log.warning("%s", warning)
    except ParseError as exc:
        if "no data rows" in str(exc).lower():
            wb.close()
            return []
        raise
    ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: dict[str, int] = {}
    for idx, value in enumerate(header_row, start=1):
        if idx < 4:
            continue
        if value is None:
            continue
        key = str(value).strip()
        if key:
            headers[key] = idx - 1

    keys = [
        "EventID",
        "StartDateTimeOffset",
        "EndDateTimeOffset",
        "DurationSeconds",
        "SystemName",
        "EventCategoryName",
        "EventDefinitionName",
        "OeeEventTypeName",
        "Notes",
    ]
    for key in keys:
        headers.setdefault(key, -1)

    records: list[DowntimeEvent] = []
    last_line_raw = ""
    row_count = 0
    bad_rows = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_count += 1
        if row_count % 50000 == 0:
            _log.debug("processed_rows=%s path=%s", row_count, source.name)
        try:
            system_idx = headers.get("SystemName", -1) + 1
            start_idx = headers.get("StartDateTimeOffset", -1) + 1
            end_idx = headers.get("EndDateTimeOffset", -1) + 1
            event_id_idx = headers.get("EventID", -1) + 1
            category_idx = headers.get("EventCategoryName", -1) + 1
            event_type_idx = headers.get("EventDefinitionName", -1) + 1
            loss_type_idx = headers.get("OeeEventTypeName", -1) + 1
            notes_idx = headers.get("Notes", -1) + 1
            duration_idx = headers.get("DurationSeconds", -1) + 1

            line_raw = _to_text(safe_cell_value(_cell(row, system_idx), row=row_idx, col=system_idx)) or last_line_raw
            if line_raw:
                last_line_raw = line_raw

            start = _to_datetime(
                safe_cell_value(_cell(row, start_idx), row=row_idx, col=start_idx)
            )
            end = _to_datetime(safe_cell_value(_cell(row, end_idx), row=row_idx, col=end_idx)) or start
            if start is None or end is None or not line_raw:
                continue

            event_id = _to_int(safe_cell_value(_cell(row, event_id_idx), row=row_idx, col=event_id_idx))
            category = _to_text(
                safe_cell_value(_cell(row, category_idx), row=row_idx, col=category_idx)
            )
            equipment_id = _normalize_equipment(category)

            event_type_raw = _to_text(
                safe_cell_value(_cell(row, event_type_idx), row=row_idx, col=event_type_idx)
            ).lower()
            loss_type_raw = _to_text(
                safe_cell_value(_cell(row, loss_type_idx), row=row_idx, col=loss_type_idx)
            ).lower()

            normalized_line_id = normalize_line_id(line_raw)
            if normalized_line_id == "line-unknown" and fallback_line_id is not None:
                normalized_line_id = fallback_line_id

            records.append(
                DowntimeEvent(
                    event_id=event_id,
                    start_time=start,
                    end_time=end,
                    duration_seconds=_to_float(
                        safe_cell_value(_cell(row, duration_idx), row=row_idx, col=duration_idx)
                    )
                    or 0.0,
                    line_id=normalized_line_id,
                    line_raw_name=line_raw,
                    equipment_id=equipment_id,
                    equipment_raw_name=category,
                    event_type=("not_scheduled" if "not scheduled" in event_type_raw else "downtime"),
                    loss_type=_normalize_loss_type(loss_type_raw),
                    is_equipment_fault=(equipment_id is not None),
                    notes=_to_text(safe_cell_value(_cell(row, notes_idx), row=row_idx, col=notes_idx)) or None,
                )
            )
        except (ValueError, TypeError, KeyError) as exc:
            bad_rows += 1
            _log.warning("Skipping Event row %s in %s: %s", row_idx, source.name, exc)
            continue

    wb.close()
    if row_count > 0 and (bad_rows / row_count) > 0.2:
        raise ParseError(f"Too many bad rows: {bad_rows}/{row_count}")
    records.sort(key=lambda x: x.start_time)
    _save_cache(cache, source_mtime, records)
    _log.info("parsed %d events from %s", len(records), source.name)
    return records


def _normalize_equipment(category: str) -> str | None:
    if not category:
        return None
    normalization = _EQUIPMENT_NORMALIZATION_OVERRIDE or EQUIPMENT_NORMALIZATION
    categories = _OPERATIONAL_CATEGORIES_OVERRIDE or OPERATIONAL_CATEGORIES
    if category in normalization:
        return normalization[category]
    norm = category.replace("(", "").replace(")", "").strip()
    for key, value in normalization.items():
        if norm.lower() == key.lower():
            return value
    if any(norm.lower().startswith(prefix) for prefix in ["caser", "wrapper"]):
        return norm.split()[0].lower()
    if norm in categories:
        return None
    return None


QUALITY_LOSS_KEYWORDS: tuple[str, ...] = (
    "quality",
    "yield",
    "scrap",
    "reject",
    "rework",
)


def _normalize_loss_type(raw: str) -> str:
    """Map raw OEE event type text to a canonical loss category.

    The caller is expected to pass *lowercased* text from the
    ``OeeEventTypeName`` column.
    """
    if "availability" in raw:
        return "availability_loss"
    if "performance" in raw:
        return "performance_loss"
    if any(kw in raw for kw in QUALITY_LOSS_KEYWORDS):
        return "quality_loss"
    if "not scheduled" in raw:
        return "system_not_scheduled"
    return "availability_loss"


def validate_event_workbook(wb: Any, path: str) -> list[str]:
    """Check expected structure before parsing. Returns warnings (empty = OK)."""
    warnings: list[str] = []
    if not getattr(wb, "sheetnames", None):
        raise ParseError(f"No sheets in {path}")
    ws = wb.active
    if ws is None or (ws.max_row is not None and ws.max_row < 1):
        raise ParseError(f"Sheet has no rows in {path}")
    if ws.max_row is not None and ws.max_row < 2:
        raise ParseError(f"Sheet has no data rows in {path}")
    max_col = int(min(int(getattr(ws, "max_column", 1) or 1), 50))
    headers = [str(ws.cell(1, col).value or "").strip().lower() for col in range(1, max_col + 1)]
    if not (set(headers) & {"line", "event", "start", "end"}):
        warnings.append(f"No expected Event columns found in {path}: got {headers[:10]}")
    return warnings


def _load_cache(cache_path: Path, source_mtime: float) -> list[DowntimeEvent] | None:
    payload = read_cache(cache_path, CACHE_VERSION)
    if payload is None:
        return None
    if float(payload.get("source_mtime", -1)) != float(source_mtime):
        return None
    records: list[DowntimeEvent] = []
    for item in payload.get("items", []):
        start = _to_datetime(item.get("start_time"))
        end = _to_datetime(item.get("end_time"))
        if start is None or end is None:
            continue
        records.append(
            DowntimeEvent(
                event_id=_to_int(item.get("event_id")),
                start_time=start,
                end_time=end,
                duration_seconds=_to_float(item.get("duration_seconds")) or 0.0,
                line_id=str(item.get("line_id", "line-unknown")),
                line_raw_name=str(item.get("line_raw_name", "")),
                equipment_id=item.get("equipment_id"),
                equipment_raw_name=str(item.get("equipment_raw_name", "")),
                event_type=str(item.get("event_type", "downtime")),
                loss_type=str(item.get("loss_type", "availability_loss")),
                is_equipment_fault=bool(item.get("is_equipment_fault", False)),
                notes=item.get("notes"),
            )
        )
    return records


def _save_cache(cache_path: Path, source_mtime: float, items: list[DowntimeEvent]) -> None:
    serializable = []
    for item in items:
        p = asdict(item)
        p["start_time"] = item.start_time.isoformat()
        p["end_time"] = item.end_time.isoformat()
        serializable.append(p)
    write_cache(cache_path, serializable, CACHE_VERSION, source_mtime)
