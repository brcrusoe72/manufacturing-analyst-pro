"""OEE Excel export parser with caching."""

from __future__ import annotations

import re
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from ._compat import ParseError, get_logger, safe_cell_value
from ._parser_utils import _cell, _infer_line_id_from_filename, _to_datetime, _to_float, _to_int, _to_text
from .utils import get_cache_dir, read_cache, write_cache

CACHE_VERSION = 5
_log = get_logger("connectors.parsers.oee")


@dataclass(slots=True)
class OEEInterval:
    """Single OEE interval extracted from MES OEE export."""

    timestamp: datetime
    line_id: str
    line_raw_name: str
    availability: float | None
    performance: float | None
    quality: float | None
    oee: float | None
    mtbf_minutes: float | None
    mttr_minutes: float | None
    total_units: int
    good_units: int
    bad_units: int
    downtime_seconds: float
    interval_seconds: float
    cases_per_hour: float


def parse_oee_file(path: str | Path, *, sheet_name: str = "Data") -> list[OEEInterval]:
    """Parse OEE Excel export into typed interval records with cache support."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for MES Excel parsing. Install with: pip install openpyxl"
        ) from exc

    source = Path(path)
    fallback_line_id = _infer_line_id_from_filename(source.name)
    cache = get_cache_dir() / f"{source.stem.replace(' ', '_').lower()}_oee.json"
    source_mtime = source.stat().st_mtime
    cached = _load_cache(cache, source_mtime)
    if cached is not None:
        _log.debug("cache hit path=%s records=%d", source.name, len(cached))
        return cached
    _log.debug("cache miss path=%s", source.name)

    wb = load_workbook(filename=source, read_only=True, data_only=True)
    try:
        for warning in validate_oee_workbook(wb, str(source)):
            _log.warning("%s", warning)
    except ParseError as exc:
        if "no data rows" in str(exc).lower():
            wb.close()
            return []
        raise
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers: dict[str, int] = {}
    for idx, value in enumerate(header_row, start=1):
        if idx < 4:
            continue
        if value is None:
            continue
        key = str(value).strip()
        if key:
            headers[key] = idx - 1

    required = [
        "GroupValue",
        "AvailabilityDecimal",
        "PerformanceDecimal",
        "QualityDecimal",
        "OeeDecimal",
        "MtbfMinutes",
        "MttrMinutes",
        "TotalDisplayUnits",
        "GoodDisplayUnits",
        "BadDisplayUnits",
        "AvailabilityLossSeconds",
        "IntervalSeconds",
    ]
    for key in required:
        if key not in headers:
            wb.close()
            raise ParseError(f"Required column {key!r} not found in headers: {list(headers.keys())}")

    records: list[OEEInterval] = []
    last_line_raw = ""
    row_count = 0
    bad_rows = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_count += 1
        if row_count % 50000 == 0:
            _log.debug("processed_rows=%s path=%s", row_count, source.name)
        try:
            line_raw = _to_text(safe_cell_value(_cell(row, 2), row=row_idx, col=2)) or last_line_raw
            if line_raw:
                last_line_raw = line_raw
            group_idx = headers["GroupValue"] + 1
            ts = _to_datetime(safe_cell_value(_cell(row, group_idx), row=row_idx, col=group_idx))
            if ts is None or not line_raw:
                continue

            availability_idx = headers["AvailabilityDecimal"] + 1
            performance_idx = headers["PerformanceDecimal"] + 1
            quality_idx = headers["QualityDecimal"] + 1
            oee_idx = headers["OeeDecimal"] + 1
            total_units_idx = headers["TotalDisplayUnits"] + 1
            good_units_idx = headers["GoodDisplayUnits"] + 1
            bad_units_idx = headers["BadDisplayUnits"] + 1
            interval_seconds_idx = headers["IntervalSeconds"] + 1
            mtbf_idx = headers["MtbfMinutes"] + 1
            mttr_idx = headers["MttrMinutes"] + 1
            downtime_idx = headers["AvailabilityLossSeconds"] + 1

            availability = _to_float(
                safe_cell_value(_cell(row, availability_idx), row=row_idx, col=availability_idx)
            )
            performance = _to_float(
                safe_cell_value(_cell(row, performance_idx), row=row_idx, col=performance_idx)
            )
            quality = _to_float(safe_cell_value(_cell(row, quality_idx), row=row_idx, col=quality_idx))
            oee = _to_float(safe_cell_value(_cell(row, oee_idx), row=row_idx, col=oee_idx))

            total_units = _to_int(
                safe_cell_value(_cell(row, total_units_idx), row=row_idx, col=total_units_idx)
            )
            good_units = _to_int(
                safe_cell_value(_cell(row, good_units_idx), row=row_idx, col=good_units_idx)
            )
            bad_units = _to_int(
                safe_cell_value(_cell(row, bad_units_idx), row=row_idx, col=bad_units_idx)
            )
            interval_seconds = (
                _to_float(
                    safe_cell_value(
                        _cell(row, interval_seconds_idx),
                        row=row_idx,
                        col=interval_seconds_idx,
                    )
                )
                or 3600.0
            )

            if (
                availability is None
                and performance is None
                and quality is None
                and oee is None
                and total_units == 0
            ):
                continue

            per_hour = 0.0
            if interval_seconds > 0:
                per_hour = float(total_units) / (interval_seconds / 3600.0)

            normalized_line_id = normalize_line_id(line_raw)
            if normalized_line_id == "line-unknown" and fallback_line_id is not None:
                normalized_line_id = fallback_line_id

            records.append(
                OEEInterval(
                    timestamp=ts,
                    line_id=normalized_line_id,
                    line_raw_name=line_raw,
                    availability=availability,
                    performance=performance,
                    quality=quality,
                    oee=oee,
                    mtbf_minutes=_to_float(
                        safe_cell_value(_cell(row, mtbf_idx), row=row_idx, col=mtbf_idx)
                    ),
                    mttr_minutes=_to_float(
                        safe_cell_value(_cell(row, mttr_idx), row=row_idx, col=mttr_idx)
                    ),
                    total_units=total_units,
                    good_units=good_units,
                    bad_units=bad_units,
                    downtime_seconds=_to_float(
                        safe_cell_value(_cell(row, downtime_idx), row=row_idx, col=downtime_idx)
                    )
                    or 0.0,
                    interval_seconds=interval_seconds,
                    cases_per_hour=per_hour,
                )
            )
        except (ValueError, TypeError, KeyError) as exc:
            bad_rows += 1
            _log.warning("Skipping OEE row %s in %s: %s", row_idx, source.name, exc)
            continue

    wb.close()
    if row_count > 0 and (bad_rows / row_count) > 0.2:
        raise ParseError(f"Too many bad rows: {bad_rows}/{row_count}")
    records.sort(key=lambda x: x.timestamp)
    _save_cache(cache, source_mtime, records)
    _log.info("parsed %d OEE records from %s", len(records), source.name)
    return records


def normalize_line_id(raw: str) -> str:
    """Normalize line naming from OEE/events/schedule into canonical id."""
    text = (raw or "").strip().lower()
    match = re.search(r"line\s*[-_ ]*(\d+)", text)
    if match:
        return f"line-{int(match.group(1))}"
    if text.isdigit():
        return f"line-{int(text)}"
    return "line-unknown"


def validate_oee_workbook(wb: Any, path: str) -> list[str]:
    """Check expected structure before parsing. Returns warnings (empty = OK)."""
    warnings: list[str] = []
    if not getattr(wb, "sheetnames", None):
        raise ParseError(f"No sheets in {path}")
    ws = wb.active
    if ws is None or (ws.max_row is not None and ws.max_row < 1):
        raise ParseError(f"Sheet has no rows in {path}")
    if ws.max_row is not None and ws.max_row < 2:
        raise ParseError(f"Sheet has no data rows in {path}")
    max_col = int(min(int(getattr(ws, "max_column", 1) or 1), 50))
    headers = [str(ws.cell(1, col).value or "").strip().lower() for col in range(1, max_col + 1)]
    if not (set(headers) & {"line", "oee"}):
        warnings.append(f"No expected OEE columns found in {path}: got {headers[:10]}")
    return warnings


def _load_cache(cache_path: Path, source_mtime: float) -> list[OEEInterval] | None:
    payload = read_cache(cache_path, CACHE_VERSION)
    if payload is None:
        return None
    if float(payload.get("source_mtime", -1)) != float(source_mtime):
        return None
    records: list[OEEInterval] = []
    for item in payload.get("items", []):
        try:
            records.append(
                OEEInterval(
                    timestamp=_to_datetime(item.get("timestamp")) or datetime.now(timezone.utc),
                    line_id=str(item.get("line_id", "line-unknown")),
                    line_raw_name=str(item.get("line_raw_name", "")),
                    availability=_to_float(item.get("availability")),
                    performance=_to_float(item.get("performance")),
                    quality=_to_float(item.get("quality")),
                    oee=_to_float(item.get("oee")),
                    mtbf_minutes=_to_float(item.get("mtbf_minutes")),
                    mttr_minutes=_to_float(item.get("mttr_minutes")),
                    total_units=_to_int(item.get("total_units")),
                    good_units=_to_int(item.get("good_units")),
                    bad_units=_to_int(item.get("bad_units")),
                    downtime_seconds=_to_float(item.get("downtime_seconds")) or 0.0,
                    interval_seconds=_to_float(item.get("interval_seconds")) or 3600.0,
                    cases_per_hour=_to_float(item.get("cases_per_hour")) or 0.0,
                )
            )
        except (ValueError, TypeError, KeyError):
            continue
    return records


def _save_cache(cache_path: Path, source_mtime: float, items: list[OEEInterval]) -> None:
    serializable = []
    for item in items:
        p = asdict(item)
        p["timestamp"] = item.timestamp.isoformat()
        serializable.append(p)
    write_cache(cache_path, serializable, CACHE_VERSION, source_mtime)
